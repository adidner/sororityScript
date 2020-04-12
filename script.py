from openpyxl import load_workbook
import smtplib, ssl

failingGirls = []
gpaMinimum = 2.0
nameGPAFile = "Base1.xlsx"
nameEmailFile = "Contact Sheet Format.xlsx"

yourEmail = ""
password = ""

port = 465
message = """\
Subject: Help with Grades

This is the Scholarship Chair for (Insert Sorority Name Here).
In accordance with our bylaws I've been asked to reach out to you because your grades have begun to slip a bit
I'd like to schedule a 1 on 1 to disscuss new tactics and anything I/the soririty can do to help you

Thanks,
Name
"""

context = ssl.create_default_context()


def grabDataFromGPATable(targetfilename):

    workbook = load_workbook(filename=targetfilename)

    sheet = workbook.active

    firstNames = sheet["A"]
    lastNames = sheet["B"]
    GPAs = sheet["E"]



    for (first,last,gpa) in zip(firstNames, lastNames, GPAs):


        if isinstance(gpa.value,float) and gpa.value <= gpaMinimum:
            newObject = {
                "first": first.value.lower(),
                "last": last.value.lower(),
                "gpa": gpa.value
            }
            failingGirls.append(newObject)

    print(failingGirls)


def grabDataFromEmailTable(targetfilename):

    workbook2 = load_workbook(filename=targetfilename)

    sheet2 = workbook2.active


    firstNames2 = sheet2["D"]
    lastNames2 = sheet2["C"]
    emails = sheet2["G"]

    print("Go here with your Gmail logged in and allow apps like this one to login as you: https://www.google.com/settings/security/lesssecureapps")

    yourEmail = input("Enter Your yourEmail: ")
    password = input("Enter your password: ")

    for currentGirl in failingGirls:
        for first, last, email in zip(firstNames2, lastNames2, emails):
            if currentGirl["first"] == first.value.lower() and currentGirl["last"] == last.value.lower():
                currentGirl["email"] = email.value
                sendEmail(currentGirl, yourEmail, password)


def sendEmail(currentGirl, yourEmail, password):

    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
        server.login(yourEmail, password)
        server.sendmail(yourEmail, yourEmail, message)







def main():

    grabDataFromGPATable(nameGPAFile)
    grabDataFromEmailTable(nameEmailFile)


if __name__ == '__main__':
    main()
